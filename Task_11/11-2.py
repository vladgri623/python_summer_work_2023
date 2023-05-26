import csv
from openpyxl import Workbook

ls = []
with open('11-2.csv', encoding='utf-8') as f:
    rows = csv.reader(f)
    for r in rows:
        ls.append(r)

lss = sorted(ls[1:], key=lambda x: x[1])
lss = sorted(lss, key=lambda x: x[2])
lss = sorted(lss, key=lambda x: x[3])
ls = ls[:1]
for i in lss:
    ls.append(i)

wb = Workbook()
ws = wb.active
for i in range(len(ls)):
    for j in range(5):
        ws.cell(row=i+1, column=j+1).value = ls[i][j]

salary_sum = 0
for i in ls[1:]:
    salary_sum += int(i[4])
ws.cell(row=ws.max_row+1, column=1).value = 'ИТОГО'
ws.cell(row=ws.max_row, column=2).value = salary_sum
wb.save('11-2.xlsx')
